"""Deterministic source-to-Markdown conversion and atomic artifact publishing."""

from __future__ import annotations

import asyncio
import base64
import csv
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from decimal import Decimal
import hashlib
from io import BytesIO
import json
import math
import mimetypes
import os
from pathlib import Path
import re
import shutil
import threading
from types import SimpleNamespace
from typing import Any, Callable, Iterable, Iterator, Mapping, Sequence
import zipfile

from markitdown import MarkItDown
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image

from app.config import (
    DEFAULT_DOCUMENT_EXCEL_ROWS_PER_PART,
    DEFAULT_DOCUMENT_TEXT_CHARS_PER_PART,
    Settings,
)
from app.llm.clients import ModelClient
from app.llm.prompts import prompt_for_client
from app.llm.registry import ModelRegistryError
from app.llm.types import ModelRole


CONVERTER_VERSION = "document-conversion-v2"
EXCEL_ROWS_PER_PART = DEFAULT_DOCUMENT_EXCEL_ROWS_PER_PART
# A bounded Markdown part is the atomic retrieval/evidence unit. 8K characters
# is typically about 2K tokens for Latin text and at most about 8K tokens for
# dense CJK text, so a 16-part answer remains viable in a 128K context window.
TEXT_CHARS_PER_PART = DEFAULT_DOCUMENT_TEXT_CHARS_PER_PART
HEARTBEAT_INTERVAL_SECONDS = 15.0
MODEL_REQUEST_TIMEOUT_SECONDS = 120.0
VISUAL_REQUEST_MAX_ATTEMPTS = 3
VISUAL_RETRY_BASE_DELAY_SECONDS = 1.0
VISUAL_CACHE_VERSION = "visual-evidence-v1"
VISUAL_MAX_OUTPUT_TOKENS = 4_096
VISUAL_REASONING_EFFORT = "low"
VISUAL_MAX_DIMENSION_PIXELS = 1_600
VISUAL_MAX_PASSTHROUGH_BYTES = 2_000_000
VISUAL_JPEG_QUALITY = 88
LEGACY_VECTOR_MEDIA_TYPES = frozenset(
    {"image/wmf", "image/x-wmf", "image/emf", "image/x-emf"}
)
LEGACY_VECTOR_FALLBACK = (
    "**Embedded visual unavailable:** a legacy WMF/EMF vector image could not "
    "be rasterized; no visual details were inferred."
)
PDF_TEXT_MIN_CHARACTERS = 40
PDF_RENDER_SCALE = 1.5
SUPPORTED_EXTENSIONS = frozenset(
    {
        ".pdf",
        ".docx",
        ".pptx",
        ".xlsx",
        ".xls",
        ".csv",
        ".tsv",
        ".txt",
        ".md",
        ".html",
        ".json",
        ".xml",
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
    }
)
IMAGE_EXTENSIONS = frozenset({".png", ".jpg", ".jpeg", ".webp"})

Heartbeat = Callable[[], None]
ProgressReporter = Callable[[Mapping[str, Any]], None]
ModelResolver = Callable[[ModelRole], ModelClient]


@contextmanager
def _continuous_heartbeat(
    heartbeat: Heartbeat,
) -> Iterator[None]:
    """Refresh a job lease while a blocking converter owns the Worker thread."""
    interval = HEARTBEAT_INTERVAL_SECONDS
    if interval <= 0:
        raise ValueError("Heartbeat interval must be greater than zero")

    stop = threading.Event()
    lock = threading.Lock()
    failures: list[BaseException] = []

    def beat() -> None:
        with lock:
            if failures:
                raise failures[0]
            heartbeat()

    def run() -> None:
        while not stop.wait(interval):
            try:
                with lock:
                    heartbeat()
            except BaseException as exc:
                with lock:
                    failures.append(exc)
                stop.set()
                return

    beat()
    thread = threading.Thread(
        target=run,
        name="document-conversion-heartbeat",
        daemon=True,
    )
    thread.start()
    try:
        yield
    finally:
        stop.set()
        thread.join()
    beat()


async def _await_with_heartbeat(
    awaitable: Any,
    heartbeat: Heartbeat,
) -> Any:
    """Await model I/O while refreshing the lease on the same event loop."""
    task = asyncio.create_task(awaitable)
    loop = asyncio.get_running_loop()
    deadline = loop.time() + MODEL_REQUEST_TIMEOUT_SECONDS
    try:
        while True:
            remaining = deadline - loop.time()
            if remaining <= 0:
                raise TimeoutError("Model request exceeded its total time limit")
            done, _pending = await asyncio.wait(
                {task},
                timeout=min(HEARTBEAT_INTERVAL_SECONDS, remaining),
            )
            if task in done:
                return await task
            heartbeat()
    finally:
        if not task.done():
            task.cancel()
            await asyncio.gather(task, return_exceptions=True)


class DocumentConversionError(RuntimeError):
    """Base expected conversion failure for one source file."""


class UnsupportedDocumentError(DocumentConversionError):
    """The source format is outside the reliable conversion set."""


class VisualConversionModelError(DocumentConversionError):
    """Visual understanding was required but its dedicated role was unusable."""


class ArtifactPublishError(DocumentConversionError):
    """A complete staged artifact could not be atomically published."""


@dataclass(frozen=True, slots=True)
class SourceDocument:
    document_id: int
    source_path: str
    source_sha256: str
    path: Path


@dataclass(frozen=True, slots=True)
class PartDraft:
    body: str
    anchors: Mapping[str, int | str]


@dataclass(frozen=True, slots=True)
class StagedArtifact:
    staging_dir: Path
    destination_dir: Path
    converted_at: datetime
    part_count: int


class DocumentConversionEngine:
    """Convert supported files without ever mutating the source path.

    MarkItDown and its OCR-aware plugin supply the general document adapters.
    Spreadsheet table bodies are deliberately handled separately with
    openpyxl/xlrd so cell values never pass through a vision model.
    """

    def __init__(
        self,
        settings: Settings,
        *,
        model_resolver: ModelResolver,
        rows_per_part: int = EXCEL_ROWS_PER_PART,
        text_chars_per_part: int = TEXT_CHARS_PER_PART,
    ) -> None:
        self.settings = settings
        self.model_resolver = model_resolver
        self.rows_per_part = rows_per_part
        self.text_chars_per_part = text_chars_per_part
        self._markitdown: MarkItDown | None = None
        self._pptx_markitdown: MarkItDown | None = None
        self._progress_reporter: ProgressReporter = lambda _progress: None
        self._progress_state: dict[str, Any] = {}
        self._last_visual_stats: dict[str, int] = {}
        self._markitdown_vision_adapter = _RoleBoundVisionAdapter(
            self._visual_enrichment
        )

    def stage(
        self,
        source: SourceDocument,
        *,
        job_id: int,
        heartbeat: Heartbeat = lambda: None,
        progress: ProgressReporter = lambda _progress: None,
    ) -> StagedArtifact:
        extension = source.path.suffix.lower()
        if extension not in SUPPORTED_EXTENSIONS:
            raise UnsupportedDocumentError(
                f"Unsupported source format: {extension or '[no extension]'}"
            )

        self._progress_reporter = progress
        self._progress_state = {
            "kind": extension.removeprefix(".") or "unknown",
            "phase": "extracting",
        }
        self._emit_progress()
        heartbeat()
        drafts = self._extract(source.path, extension, heartbeat)
        drafts = self._bound_drafts(drafts)
        if not drafts:
            raise DocumentConversionError("The converter produced no Markdown parts")
        self._emit_progress(phase="extracted", total_parts=len(drafts))

        converted_at = datetime.now(timezone.utc)
        staging_dir = self._staging_dir(job_id, source.document_id)
        destination_dir = self._destination_dir(source.document_id)
        self._prepare_staging_dir(staging_dir)

        manifest_parts: list[dict[str, Any]] = []
        self._emit_progress(phase="writing", total_parts=len(drafts), written_parts=0)
        try:
            for number, draft in enumerate(drafts, start=1):
                part_id = f"part-{number:03d}"
                filename = f"{part_id}.md"
                markdown = _render_part(
                    source,
                    part_id=part_id,
                    converted_at=converted_at,
                    anchors=draft.anchors,
                    body=draft.body,
                )
                part_path = staging_dir / filename
                _write_fsynced(part_path, markdown.encode("utf-8"))
                manifest_parts.append(
                    {
                        "part_id": part_id,
                        "path": filename,
                        "anchors": dict(draft.anchors),
                        "sha256": hashlib.sha256(markdown.encode("utf-8")).hexdigest(),
                    }
                )
                if number == len(drafts) or number % 25 == 0:
                    self._emit_progress(written_parts=number)

            manifest = {
                "document_id": source.document_id,
                "source_path": source.source_path,
                "source_sha256": source.source_sha256,
                "converted_at": converted_at.isoformat(),
                "converter_version": CONVERTER_VERSION,
                "status": "READY",
                "parts": manifest_parts,
            }
            manifest_bytes = (
                json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True)
                + "\n"
            ).encode("utf-8")
            _write_fsynced(staging_dir / "manifest.json", manifest_bytes)
            _fsync_directory(staging_dir)
        except BaseException:
            shutil.rmtree(staging_dir, ignore_errors=True)
            raise

        heartbeat()
        self._emit_progress(phase="publishing", written_parts=len(drafts))
        return StagedArtifact(
            staging_dir=staging_dir,
            destination_dir=destination_dir,
            converted_at=converted_at,
            part_count=len(drafts),
        )

    def _emit_progress(self, **updates: Any) -> None:
        self._progress_state.update(updates)
        self._progress_state["updated_at"] = datetime.now(timezone.utc).isoformat()
        self._progress_reporter(dict(self._progress_state))

    def publish(self, artifact: StagedArtifact) -> None:
        """Publish a complete directory and retain the last good one on error."""
        destination = artifact.destination_dir
        staging = artifact.staging_dir
        destination.parent.mkdir(parents=True, exist_ok=True)
        _assert_generated_path(destination.parent, destination)
        _assert_generated_path(staging.parent, staging)

        backup = staging.parent / f".{destination.name}.previous"
        if os.path.lexists(backup):
            if backup.is_symlink() or not backup.is_dir():
                raise ArtifactPublishError("Unsafe previous-artifact backup path")
            shutil.rmtree(backup)

        moved_previous = False
        try:
            if os.path.lexists(destination):
                if destination.is_symlink() or not destination.is_dir():
                    raise ArtifactPublishError("Unsafe Markdown artifact destination")
                os.replace(destination, backup)
                moved_previous = True
            os.replace(staging, destination)
            _fsync_directory(destination.parent)
        except Exception as exc:
            if moved_previous and os.path.lexists(backup):
                try:
                    if os.path.lexists(destination):
                        shutil.rmtree(destination, ignore_errors=True)
                    os.replace(backup, destination)
                    _fsync_directory(destination.parent)
                except Exception as rollback_exc:
                    raise ArtifactPublishError(
                        "Markdown artifact publish failed and the previous artifact "
                        "could not be restored"
                    ) from rollback_exc
            raise ArtifactPublishError(
                "Markdown artifact publish failed; the previous artifact was preserved"
            ) from exc
        else:
            if moved_previous:
                shutil.rmtree(backup, ignore_errors=True)

    def discard(self, artifact: StagedArtifact) -> None:
        shutil.rmtree(artifact.staging_dir, ignore_errors=True)

    def _extract(
        self,
        path: Path,
        extension: str,
        heartbeat: Heartbeat,
    ) -> list[PartDraft]:
        if extension == ".xlsx":
            return list(self._extract_xlsx(path, heartbeat))
        if extension == ".xls":
            return list(self._extract_xls(path, heartbeat))
        if extension in {".csv", ".tsv"}:
            return list(
                self._extract_delimited(
                    path,
                    delimiter="," if extension == ".csv" else "\t",
                    heartbeat=heartbeat,
                )
            )
        if extension in IMAGE_EXTENSIONS:
            return [self._extract_image(path, heartbeat)]
        if extension == ".pdf":
            return self._extract_pdf(path, heartbeat)
        if extension == ".pptx":
            return self._extract_pptx(path, heartbeat)
        if extension in {".txt", ".md"}:
            text = _read_text(path)
            self._emit_progress(deterministic_text_characters=len(text))
            return self._chunk_sections(text, section_prefix="section")
        if extension == ".json":
            return self._extract_json(path)
        if extension == ".xml":
            return self._extract_xml(path)

        markdown = self._convert_with_markitdown(path, heartbeat=heartbeat)
        return self._chunk_sections(markdown, section_prefix="section")

    @property
    def markitdown(self) -> MarkItDown:
        if self._markitdown is None:
            # The OCR plugin receives only a local facade whose implementation
            # delegates to the application's DOCUMENT_CONVERSION role. It
            # cannot own or select a Provider/Profile/model itself.
            self._markitdown = MarkItDown(
                enable_plugins=True,
                llm_client=self._markitdown_vision_adapter,
                # This is only the plugin's required label. The adapter never
                # sends it remotely; the role-resolved ModelClient owns the
                # actual Profile/model name.
                llm_model=ModelRole.DOCUMENT_CONVERSION.value,
            )
        return self._markitdown

    @property
    def pptx_markitdown(self) -> MarkItDown:
        if self._pptx_markitdown is None:
            # The OCR plugin silently falls back to CPU-heavy local ONNX OCR
            # when a vision request fails. That fallback can monopolize the
            # Worker process and hide the real Provider error. PPTX uses the
            # built-in converter with the same role-bound vision adapter so a
            # failed model request remains explicit and retryable.
            self._pptx_markitdown = MarkItDown(
                enable_plugins=False,
                llm_client=self._markitdown_vision_adapter,
                llm_model=ModelRole.DOCUMENT_CONVERSION.value,
            )
        return self._pptx_markitdown

    def _convert_with_markitdown(
        self,
        path: Path,
        *,
        allow_empty: bool = False,
        heartbeat: Heartbeat | None = None,
    ) -> str:
        return self._convert_using_markitdown(
            self.markitdown,
            path,
            allow_empty=allow_empty,
            heartbeat=heartbeat,
        )

    def _convert_pptx_with_markitdown(
        self,
        path: Path,
        *,
        heartbeat: Heartbeat,
    ) -> str:
        return self._convert_using_markitdown(
            self.pptx_markitdown,
            path,
            heartbeat=heartbeat,
        )

    def _convert_using_markitdown(
        self,
        converter: MarkItDown,
        path: Path,
        *,
        allow_empty: bool = False,
        heartbeat: Heartbeat | None = None,
    ) -> str:
        self._markitdown_vision_adapter.reset(
            heartbeat=heartbeat,
            collect_visuals=True,
        )
        try:
            result = converter.convert_local(path, file_extension=path.suffix)
        except Exception as exc:
            self._markitdown_vision_adapter.raise_if_failed()
            raise DocumentConversionError(
                f"Deterministic extraction failed for {path.suffix.lower()}"
            ) from exc
        self._markitdown_vision_adapter.raise_if_failed()
        markdown = result.markdown
        collected = self._markitdown_vision_adapter.collected_visuals
        progress_updates: dict[str, Any] = {
            "phase": "visual_enrichment" if collected else "extracted",
            "embedded_visuals_total": len(collected),
            "embedded_visuals_completed": 0,
            "embedded_visuals_cache_hits": 0,
            "embedded_visuals_model_requests": 0,
            "embedded_visuals_model_completed": 0,
            "embedded_visuals_legacy": 0,
            "deterministic_text_characters": len(_normalize_markdown(markdown)),
        }
        if path.suffix.lower() == ".pptx":
            slide_pattern = re.compile(
                r"<!--\s*Slide\s+number:\s*(\d+)\s*-->",
                flags=re.IGNORECASE,
            )
            slide_markers = list(slide_pattern.finditer(markdown))
            slide_numbers = [int(marker.group(1)) for marker in slide_markers]
            progress_updates["total_slides"] = max(slide_numbers, default=1)
            progress_updates["slides_extracted"] = max(slide_numbers, default=1)
            slide_blocks = [
                markdown[
                    marker.start() : (
                        slide_markers[index + 1].start()
                        if index + 1 < len(slide_markers)
                        else len(markdown)
                    )
                ]
                for index, marker in enumerate(slide_markers)
            ] or [markdown]
            progress_updates["slides_with_visuals"] = sum(
                any(placeholder in block for placeholder, _, _ in collected)
                for block in slide_blocks
            )
            progress_updates["slides_with_text"] = sum(
                bool(
                    _normalize_markdown(
                        slide_pattern.sub(
                            "",
                            re.sub(
                                r"<!--\s*(?:mini-)?kb-visual-[^>]+-->",
                                "",
                                block,
                                flags=re.IGNORECASE,
                            ),
                        )
                    )
                )
                for block in slide_blocks
            )
        self._emit_progress(**progress_updates)
        if collected:
            descriptions = self._visual_enrichments(
                [(image_bytes, media_type) for _, image_bytes, media_type in collected],
                heartbeat or (lambda: None),
                progress_scope="embedded_visuals",
            )
            for (placeholder, _, _), description in zip(
                collected,
                descriptions,
                strict=True,
            ):
                markdown = markdown.replace(placeholder, description)
            self._emit_progress(phase="extracted")
        markdown = _normalize_markdown(markdown)
        if not markdown and not allow_empty:
            raise DocumentConversionError("Deterministic extraction produced no text")
        return markdown

    def _extract_pdf(self, path: Path, heartbeat: Heartbeat) -> list[PartDraft]:
        try:
            import pymupdf

            drafts: list[PartDraft | None] = []
            visual_pages: list[tuple[int, int, str]] = []
            batch_size = max(1, self.settings.document_visual_concurrency * 2)

            with pymupdf.open(path) as document:
                total_pages = document.page_count
                direct_text_pages = 0
                self._emit_progress(
                    phase="analyzing",
                    total_pages=total_pages,
                    analyzed_pages=0,
                    direct_text_pages=0,
                    visual_pages=0,
                    visual_pages_completed=0,
                    visual_cache_hits=0,
                    model_requests=0,
                )
                for page_number, page in enumerate(document, start=1):
                    heartbeat()
                    text = _normalize_markdown(page.get_text("text", sort=True) or "")
                    if _has_meaningful_pdf_text(text):
                        direct_text_pages += 1
                        drafts.append(
                            PartDraft(
                                body=f"## Page {page_number}\n\n{text}",
                                anchors={"page": page_number},
                            )
                        )
                    else:
                        drafts.append(None)
                        visual_pages.append(
                            (len(drafts) - 1, page_number, text)
                        )
                    if page_number == total_pages or page_number % 25 == 0:
                        self._emit_progress(
                            analyzed_pages=page_number,
                            direct_text_pages=direct_text_pages,
                            visual_pages=len(visual_pages),
                        )

                self._emit_progress(
                    phase="visual_enrichment" if visual_pages else "extracted",
                    analyzed_pages=total_pages,
                    direct_text_pages=direct_text_pages,
                    visual_pages=len(visual_pages),
                )
                visual_completed = 0
                visual_cache_hits = 0
                model_requests = 0
                for start in range(0, len(visual_pages), batch_size):
                    heartbeat()
                    batch = visual_pages[start : start + batch_size]
                    images: list[tuple[bytes, str]] = []
                    for _, page_number, _ in batch:
                        page = document.load_page(page_number - 1)
                        pixmap = page.get_pixmap(
                            matrix=pymupdf.Matrix(PDF_RENDER_SCALE, PDF_RENDER_SCALE),
                            alpha=False,
                        )
                        images.append((pixmap.tobytes("png"), "image/png"))

                    descriptions = self._visual_enrichments(images, heartbeat)
                    stats = self._last_visual_stats
                    visual_completed += len(batch)
                    visual_cache_hits += int(stats.get("cache_hit_items", 0))
                    model_requests += int(stats.get("model_requests", 0))
                    for (draft_index, page_number, extracted_text), description in zip(
                        batch,
                        descriptions,
                        strict=True,
                    ):
                        text_section = (
                            f"### Deterministic text\n\n{extracted_text}\n\n"
                            if extracted_text
                            else ""
                        )
                        drafts[draft_index] = PartDraft(
                            body=(
                                f"## Page {page_number}\n\n"
                                f"{text_section}"
                                "### Visual extraction\n\n"
                                f"{description}"
                            ),
                            anchors={"page": page_number},
                        )
                    self._emit_progress(
                        visual_pages_completed=visual_completed,
                        visual_cache_hits=visual_cache_hits,
                        model_requests=model_requests,
                    )
                self._emit_progress(phase="extracted")
            return [draft for draft in drafts if draft is not None]
        except DocumentConversionError:
            raise
        except BaseException as exc:
            if not isinstance(exc, Exception):
                raise
            raise DocumentConversionError("PDF page extraction failed") from exc

    def _describe_pdf_page(
        self,
        path: Path,
        page_number: int,
        heartbeat: Heartbeat,
    ) -> str:
        try:
            import pymupdf

            with pymupdf.open(path) as document:
                page = document.load_page(page_number - 1)
                pixmap = page.get_pixmap(
                    matrix=pymupdf.Matrix(2, 2),
                    alpha=False,
                )
                image_bytes = pixmap.tobytes("png")
                dimensions = f"{page.rect.width:g} x {page.rect.height:g} points"
        except Exception as exc:
            raise DocumentConversionError(
                f"Scanned PDF page {page_number} could not be rendered"
            ) from exc

        heartbeat()
        description = self._visual_enrichment(
            image_bytes,
            "image/png",
            heartbeat,
        )
        heartbeat()
        return (
            f"## Page {page_number}\n\n"
            f"Page dimensions: {dimensions}\n\n"
            "### Visual extraction\n\n"
            f"{description}"
        )

    def _extract_pptx(
        self,
        path: Path,
        heartbeat: Heartbeat,
    ) -> list[PartDraft]:
        # PPTX conversion is one blocking MarkItDown call that may make many
        # sequential vision requests for embedded images. Keep the persisted
        # job lease fresh throughout that call so the Web recovery path cannot
        # mistake a healthy Worker for a crashed one.
        with _continuous_heartbeat(heartbeat):
            markdown = self._convert_pptx_with_markitdown(
                path,
                heartbeat=heartbeat,
            )
        # markitdown-ocr 0.1 emits escaped line separators in its PPTX
        # adapter. Normalize those adapter separators before part splitting.
        if "<!-- Slide number:" in markdown:
            markdown = markdown.replace("\\n", "\n")
        split = _split_anchored_blocks(
            markdown,
            re.compile(r"<!--\s*Slide\s+number:\s*(\d+)\s*-->", re.IGNORECASE),
            anchor_name="slide",
        )
        if split:
            return split
        return [PartDraft(body=markdown, anchors={"slide": 1})]

    def _extract_xlsx(
        self,
        path: Path,
        heartbeat: Heartbeat,
    ) -> Iterator[PartDraft]:
        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise DocumentConversionError("XLSX workbook could not be opened") from exc

        try:
            yielded = False
            total_sheets = len(workbook.worksheets)
            completed_sheets = 0
            table_parts = 0
            self._emit_progress(
                phase="table_extraction",
                total_sheets=total_sheets,
                completed_sheets=0,
                table_parts_completed=0,
            )
            for worksheet in workbook.worksheets:
                heartbeat()
                sheet_yielded = False
                rows = (
                    (row_number, tuple(values))
                    for row_number, values in enumerate(
                        worksheet.iter_rows(values_only=True),
                        start=1,
                    )
                )
                for draft in self._table_parts(
                    rows,
                    sheet_name=worksheet.title,
                ):
                    yielded = sheet_yielded = True
                    table_parts += 1
                    yield draft
                    heartbeat()
                if not sheet_yielded:
                    yielded = True
                    table_parts += 1
                    yield PartDraft(
                        body=f"## Sheet: {worksheet.title}\n\n_Empty sheet._",
                        anchors={"sheet": worksheet.title, "rows": "empty"},
                    )
                completed_sheets += 1
                self._emit_progress(
                    completed_sheets=completed_sheets,
                    table_parts_completed=table_parts,
                )
            if not yielded:
                raise DocumentConversionError("XLSX workbook contains no sheets")
        finally:
            workbook.close()

        if _xlsx_contains_embedded_media(path):
            yield from self._extract_xlsx_images(path, heartbeat)
        else:
            self._emit_progress(phase="extracted")

    def _extract_xlsx_images(
        self,
        path: Path,
        heartbeat: Heartbeat,
    ) -> Iterator[PartDraft]:
        try:
            workbook = load_workbook(
                filename=path,
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise DocumentConversionError(
                "XLSX embedded images could not be opened"
            ) from exc

        try:
            records: list[dict[str, Any]] = []
            for worksheet in workbook.worksheets:
                for image_number, image in enumerate(
                    getattr(worksheet, "_images", ()),
                    start=1,
                ):
                    try:
                        image_bytes = image._data()
                    except Exception as exc:
                        raise DocumentConversionError(
                            "XLSX embedded image data could not be read"
                        ) from exc

                    try:
                        anchor = getattr(image, "anchor", None)
                        marker = getattr(anchor, "_from", None)
                        row_number = int(marker.row) + 1 if marker is not None else 1
                        column_number = int(marker.col) + 1 if marker is not None else 1
                        cell = f"{get_column_letter(column_number)}{row_number}"
                    except Exception:
                        row_number = 1
                        cell = "unknown"

                    try:
                        with Image.open(BytesIO(image_bytes)) as decoded:
                            decoded.load()
                            width, height = decoded.size
                            image_format = decoded.format or "unknown"
                            media_type = Image.MIME.get(decoded.format or "") or "image/png"
                    except Exception as exc:
                        raise DocumentConversionError(
                            "XLSX embedded image could not be decoded locally"
                        ) from exc

                    records.append(
                        {
                            "image_number": image_number,
                            "sheet": worksheet.title,
                            "row_number": row_number,
                            "cell": cell,
                            "image_format": image_format,
                            "width": width,
                            "height": height,
                            "image_bytes": image_bytes,
                            "media_type": media_type,
                        }
                    )

            self._emit_progress(
                phase="visual_enrichment",
                embedded_visuals_total=len(records),
                embedded_visuals_completed=0,
            )
            descriptions = self._visual_enrichments(
                [
                    (record["image_bytes"], record["media_type"])
                    for record in records
                ],
                heartbeat,
                progress_scope="embedded_visuals",
            )
            for record, description in zip(records, descriptions, strict=True):
                heartbeat()
                yield PartDraft(
                    body=(
                        f"## Sheet image {record['image_number']}\n\n"
                        f"- Sheet: {record['sheet']}\n"
                        f"- Anchor cell: {record['cell']}\n"
                        f"- Format: {record['image_format']}\n"
                        f"- Dimensions: {record['width']} x {record['height']} pixels\n\n"
                        "### Visual extraction\n\n"
                        f"{description}"
                    ),
                    anchors={
                        "sheet": record["sheet"],
                        "rows": f"{record['row_number']}-{record['row_number']}",
                    },
                )
            self._emit_progress(phase="extracted")
        finally:
            workbook.close()

    def _extract_xls(
        self,
        path: Path,
        heartbeat: Heartbeat,
    ) -> Iterator[PartDraft]:
        try:
            import xlrd

            workbook = xlrd.open_workbook(path, on_demand=True)
        except Exception as exc:
            raise DocumentConversionError("XLS workbook could not be opened") from exc

        try:
            sheets = workbook.sheets()
            self._emit_progress(
                phase="table_extraction",
                total_sheets=len(sheets),
                completed_sheets=0,
                table_parts_completed=0,
            )
            table_parts = 0
            for sheet_number, sheet in enumerate(sheets, start=1):
                heartbeat()

                def rows() -> Iterator[tuple[int, tuple[Any, ...]]]:
                    for row_index in range(sheet.nrows):
                        values = tuple(
                            _xls_cell_value(sheet.cell(row_index, column), workbook)
                            for column in range(sheet.ncols)
                        )
                        yield row_index + 1, values

                sheet_yielded = False
                for draft in self._table_parts(rows(), sheet_name=sheet.name):
                    sheet_yielded = True
                    table_parts += 1
                    yield draft
                    heartbeat()
                if not sheet_yielded:
                    table_parts += 1
                    yield PartDraft(
                        body=f"## Sheet: {sheet.name}\n\n_Empty sheet._",
                        anchors={"sheet": sheet.name, "rows": "empty"},
                    )
                self._emit_progress(
                    completed_sheets=sheet_number,
                    table_parts_completed=table_parts,
                )
            self._emit_progress(phase="extracted")
        finally:
            workbook.release_resources()

    def _extract_delimited(
        self,
        path: Path,
        *,
        delimiter: str,
        heartbeat: Heartbeat,
    ) -> Iterator[PartDraft]:
        encoding = _detect_text_encoding(path)
        try:
            self._emit_progress(
                phase="table_extraction",
                table_parts_completed=0,
            )
            with path.open("r", encoding=encoding, newline="") as source_file:
                reader = csv.reader(source_file, delimiter=delimiter)
                rows = (
                    (row_number, tuple(row))
                    for row_number, row in enumerate(reader, start=1)
                )
                yielded = False
                table_parts = 0
                for draft in self._table_parts(rows, sheet_name=None):
                    yielded = True
                    table_parts += 1
                    yield draft
                    self._emit_progress(table_parts_completed=table_parts)
                    heartbeat()
                if not yielded:
                    yield PartDraft(
                        body="_Empty delimited file._",
                        anchors={"rows": "empty"},
                    )
                self._emit_progress(phase="extracted")
        except (UnicodeError, csv.Error, OSError) as exc:
            raise DocumentConversionError("Delimited text extraction failed") from exc

    def _table_parts(
        self,
        rows: Iterable[tuple[int, Sequence[Any]]],
        *,
        sheet_name: str | None,
    ) -> Iterator[PartDraft]:
        batch: list[tuple[int, Sequence[Any]]] = []
        batch_start: int | None = None

        for row_number, values in rows:
            if batch_start is None:
                batch_start = row_number
            batch.append((row_number, values))
            if row_number - batch_start + 1 >= self.rows_per_part:
                draft = _render_table_batch(batch, sheet_name=sheet_name)
                if draft is not None:
                    yield draft
                batch = []
                batch_start = None

        if batch:
            draft = _render_table_batch(batch, sheet_name=sheet_name)
            if draft is not None:
                yield draft

    def _extract_image(self, path: Path, heartbeat: Heartbeat) -> PartDraft:
        try:
            with Image.open(path) as image:
                image.load()
                width, height = image.size
                detected_format = image.format or path.suffix.removeprefix(".").upper()
                mode = image.mode
                frame_count = getattr(image, "n_frames", 1)
        except Exception as exc:
            raise DocumentConversionError("Image could not be decoded locally") from exc

        image_bytes = path.read_bytes()
        media_type = mimetypes.guess_type(path.name)[0] or "image/png"
        self._emit_progress(
            phase="visual_enrichment",
            image_width=width,
            image_height=height,
            image_format=detected_format,
            image_frames=frame_count,
        )
        description = self._visual_enrichment(
            image_bytes,
            media_type,
            heartbeat,
            progress_scope="image",
        )
        self._emit_progress(
            phase="extracted",
            image_width=width,
            image_height=height,
            image_format=detected_format,
            image_frames=frame_count,
        )
        body = (
            "# Image\n\n"
            f"- Format: {detected_format}\n"
            f"- Dimensions: {width} x {height} pixels\n"
            f"- Color mode: {mode}\n"
            f"- Frames: {frame_count}\n\n"
            "## Visual extraction\n\n"
            f"{description}"
        )
        return PartDraft(body=body, anchors={"section": "image"})

    def _visual_enrichment(
        self,
        image_bytes: bytes,
        media_type: str,
        heartbeat: Heartbeat = lambda: None,
        progress_scope: str | None = None,
    ) -> str:
        return self._visual_enrichments(
            [(image_bytes, media_type)],
            heartbeat,
            progress_scope=progress_scope,
        )[0]

    def _visual_enrichments(
        self,
        images: Sequence[tuple[bytes, str]],
        heartbeat: Heartbeat = lambda: None,
        *,
        progress_scope: str | None = None,
    ) -> list[str]:
        if not images:
            self._last_visual_stats = {
                "total_items": 0,
                "cache_hit_items": 0,
                "model_requests": 0,
            }
            return []
        prepared_images: list[tuple[int, bytes, str]] = []
        result_slots: list[str | None] = [None] * len(images)
        legacy_items = 0
        for index, (image_bytes, media_type) in enumerate(images):
            prepared = _prepare_visual_input(image_bytes, media_type)
            if prepared is None:
                result_slots[index] = LEGACY_VECTOR_FALLBACK
                legacy_items += 1
                continue
            prepared_images.append((index, *prepared))
        if not prepared_images:
            self._last_visual_stats = {
                "total_items": len(images),
                "cache_hit_items": 0,
                "model_requests": 0,
                "legacy_items": legacy_items,
            }
            return [result for result in result_slots if result is not None]
        try:
            client = self.model_resolver(ModelRole.DOCUMENT_CONVERSION)
        except ModelRegistryError as exc:
            raise VisualConversionModelError(
                "Visual understanding is required, but ModelRole."
                "DOCUMENT_CONVERSION is not configured with a usable vision model"
            ) from exc
        prompt = prompt_for_client(
            client,
            ModelRole.DOCUMENT_CONVERSION,
            "visual_evidence",
        )
        profile = getattr(client, "profile", None)
        model_cache_key = ":".join(
            (
                str(getattr(profile, "id", "")),
                str(
                    getattr(
                        profile,
                        "remote_model_name",
                        client.__class__.__qualname__,
                    )
                ),
            )
        )
        cache_hit_items = 0
        pending_by_digest: dict[
            str,
            tuple[list[int], Path, bytes, str],
        ] = {}
        for index, image, media in prepared_images:
            digest = hashlib.sha256(
                "\0".join(
                    (
                        VISUAL_CACHE_VERSION,
                        model_cache_key,
                        prompt,
                        media,
                        str(VISUAL_MAX_OUTPUT_TOKENS),
                        VISUAL_REASONING_EFFORT,
                    )
                ).encode("utf-8")
                + b"\0"
                + image
            ).hexdigest()
            cache_path = _visual_cache_path(self.settings.markdown_dir, digest)
            cached = _read_visual_cache(cache_path)
            if cached:
                result_slots[index] = cached
                cache_hit_items += 1
                continue
            existing = pending_by_digest.get(digest)
            if existing is None:
                pending_by_digest[digest] = ([index], cache_path, image, media)
            else:
                existing[0].append(index)

        pending_requests = list(pending_by_digest.values())
        self._last_visual_stats = {
            "total_items": len(images),
            "cache_hit_items": cache_hit_items,
            "model_requests": len(pending_requests),
            "legacy_items": legacy_items,
        }
        scope_completed = cache_hit_items + legacy_items
        scope_model_completed = 0
        if progress_scope:
            self._emit_progress(
                **{
                    f"{progress_scope}_total": len(images),
                    f"{progress_scope}_completed": scope_completed,
                    f"{progress_scope}_cache_hits": cache_hit_items,
                    f"{progress_scope}_model_requests": len(pending_requests),
                    f"{progress_scope}_model_completed": 0,
                    f"{progress_scope}_legacy": legacy_items,
                }
            )
        if not pending_requests:
            return [result for result in result_slots if result is not None]

        async def generate_all() -> list[Any]:
            semaphore = asyncio.Semaphore(
                self.settings.document_visual_concurrency
            )

            async def generate_one(
                indices: list[int],
                cache_path: Path,
                image: bytes,
                image_media_type: str,
            ) -> Any:
                nonlocal scope_completed, scope_model_completed
                async with semaphore:
                    generated = None
                    for attempt in range(VISUAL_REQUEST_MAX_ATTEMPTS):
                        try:
                            generated = await _await_with_heartbeat(
                                client.generate_multimodal(
                                    prompt,
                                    image,
                                    image_media_type=image_media_type,
                                    max_output_tokens=VISUAL_MAX_OUTPUT_TOKENS,
                                    reasoning_effort=VISUAL_REASONING_EFFORT,
                                ),
                                heartbeat,
                            )
                            break
                        except Exception:
                            if attempt + 1 >= VISUAL_REQUEST_MAX_ATTEMPTS:
                                raise
                            heartbeat()
                            await asyncio.sleep(
                                VISUAL_RETRY_BASE_DELAY_SECONDS * (2**attempt)
                            )
                    if generated is None:
                        raise AssertionError("visual request retry loop did not return")

                text = _normalize_markdown(generated.text)
                if not text:
                    raise VisualConversionModelError(
                        "The document_conversion vision model returned no usable content"
                    )
                for index in indices:
                    result_slots[index] = text
                _write_visual_cache(cache_path, text)
                if progress_scope:
                    scope_completed += len(indices)
                    scope_model_completed += 1
                    self._emit_progress(
                        **{
                            f"{progress_scope}_completed": scope_completed,
                            f"{progress_scope}_model_completed": scope_model_completed,
                        }
                    )
                return generated

            return list(
                await asyncio.gather(
                    *(
                        generate_one(indices, cache_path, image, media)
                        for indices, cache_path, image, media in pending_requests
                    ),
                    return_exceptions=True,
                )
            )

        generated = asyncio.run(generate_all())
        failures: list[Exception] = []
        control_signal: BaseException | None = None
        for (_indices, _cache_path, _, _), item in zip(
            pending_requests,
            generated,
            strict=True,
        ):
            if isinstance(item, BaseException):
                if isinstance(item, Exception):
                    failures.append(item)
                else:
                    control_signal = item
                continue

        if control_signal is not None:
            raise control_signal
        if failures:
            failure = failures[0]
            cause_name = type(failure.__cause__ or failure).__name__
            raise VisualConversionModelError(
                "The document_conversion vision model failed during image enrichment "
                f"({type(failure).__name__}; cause: {cause_name})"
            ) from failure
        if any(not text for text in result_slots):
            raise VisualConversionModelError(
                "The document_conversion vision model returned no usable content"
            )
        return [result for result in result_slots if result is not None]

    def _extract_json(self, path: Path) -> list[PartDraft]:
        try:
            value = json.loads(_read_text(path))
        except (json.JSONDecodeError, UnicodeError, OSError) as exc:
            raise DocumentConversionError("JSON source is not valid JSON") from exc
        normalized = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)
        self._emit_progress(deterministic_text_characters=len(normalized))
        return [
            PartDraft(
                body=f"```json\n{normalized}\n```",
                anchors={"section": "document"},
            )
        ]

    def _extract_xml(self, path: Path) -> list[PartDraft]:
        try:
            from defusedxml import ElementTree

            ElementTree.parse(path)
            source_text = _read_text(path)
        except Exception as exc:
            raise DocumentConversionError("XML source is not valid XML") from exc
        self._emit_progress(deterministic_text_characters=len(source_text))
        return [
            PartDraft(
                body=f"```xml\n{source_text.strip()}\n```",
                anchors={"section": "document"},
            )
        ]

    def _chunk_sections(self, text: str, *, section_prefix: str) -> list[PartDraft]:
        normalized = _normalize_markdown(text)
        if not normalized:
            raise DocumentConversionError("Deterministic extraction produced no text")
        chunks = _chunk_text(normalized, self.text_chars_per_part)
        return [
            PartDraft(
                body=chunk,
                anchors={"section": f"{section_prefix}-{number}"},
            )
            for number, chunk in enumerate(chunks, start=1)
        ]

    def _bound_drafts(self, drafts: Sequence[PartDraft]) -> list[PartDraft]:
        """Split any oversized page/slide/table without losing its source anchor."""
        bounded: list[PartDraft] = []
        for draft in drafts:
            chunks = _chunk_markdown(
                _normalize_markdown(draft.body),
                self.text_chars_per_part,
            )
            if len(chunks) == 1:
                bounded.append(draft)
                continue
            total = len(chunks)
            for number, chunk in enumerate(chunks, start=1):
                bounded.append(
                    PartDraft(
                        body=chunk,
                        anchors={
                            **dict(draft.anchors),
                            "segment": f"{number}/{total}",
                        },
                    )
                )
        return bounded

    def _staging_dir(self, job_id: int, document_id: int) -> Path:
        if job_id <= 0 or document_id <= 0:
            raise DocumentConversionError("Job and document ids must be positive")
        return (
            self.settings.markdown_dir
            / ".staging"
            / str(job_id)
            / str(document_id)
        )

    def _destination_dir(self, document_id: int) -> Path:
        if document_id <= 0:
            raise DocumentConversionError("Document id must be positive")
        return self.settings.markdown_dir / str(document_id)

    def _prepare_staging_dir(self, staging_dir: Path) -> None:
        staging_dir.parent.mkdir(parents=True, exist_ok=True)
        _assert_generated_path(self.settings.markdown_dir, staging_dir)
        if os.path.lexists(staging_dir):
            if staging_dir.is_symlink() or not staging_dir.is_dir():
                raise ArtifactPublishError("Unsafe Markdown staging path")
            shutil.rmtree(staging_dir)
        staging_dir.mkdir()


class _RoleBoundVisionAdapter:
    """Small sync facade used only by markitdown-ocr's plugin interface.

    The plugin expects an OpenAI-shaped ``chat.completions.create`` object.
    This facade decodes that local call and delegates to the engine's
    role-resolved ModelClient path. It owns no Provider, credential, Profile,
    or remote model name.
    """

    def __init__(
        self,
        enrich: Callable[[bytes, str, Heartbeat], str],
    ) -> None:
        self._enrich = enrich
        self._failures: list[DocumentConversionError] = []
        self._heartbeat: Heartbeat = lambda: None
        self._collect_visuals = False
        self._collected_visuals: list[tuple[str, bytes, str]] = []
        self.chat = SimpleNamespace(
            completions=SimpleNamespace(create=self._create),
        )

    def reset(
        self,
        *,
        heartbeat: Heartbeat | None = None,
        collect_visuals: bool = False,
    ) -> None:
        self._failures.clear()
        self._heartbeat = heartbeat or (lambda: None)
        self._collect_visuals = collect_visuals
        self._collected_visuals.clear()

    @property
    def collected_visuals(self) -> list[tuple[str, bytes, str]]:
        return list(self._collected_visuals)

    def raise_if_failed(self) -> None:
        if self._failures:
            raise self._failures[0]

    def _create(self, *, messages: Sequence[Mapping[str, Any]], **_kwargs: Any):
        try:
            self._heartbeat()
            prompt, image_bytes, media_type = _decode_plugin_vision_message(messages)
            if self._collect_visuals:
                text = f"<!-- mini-kb-visual-{len(self._collected_visuals) + 1:04d} -->"
                self._collected_visuals.append((text, image_bytes, media_type))
            else:
                text = self._enrich_with_prompt(image_bytes, media_type, prompt)
            self._heartbeat()
        except DocumentConversionError as exc:
            self._failures.append(exc)
            raise
        except Exception as exc:
            failure = VisualConversionModelError(
                "markitdown-ocr could not prepare embedded visual content"
            )
            self._failures.append(failure)
            raise failure from exc
        return SimpleNamespace(
            choices=[SimpleNamespace(message=SimpleNamespace(content=text))]
        )

    def _enrich_with_prompt(
        self,
        image_bytes: bytes,
        media_type: str,
        _plugin_prompt: str,
    ) -> str:
        # Use the application's stricter evidence prompt rather than allowing
        # a plugin prompt to become a second model-policy boundary.
        return self._enrich(image_bytes, media_type, self._heartbeat)


def _decode_plugin_vision_message(
    messages: Sequence[Mapping[str, Any]],
) -> tuple[str, bytes, str]:
    prompt = ""
    data_uri = ""
    for message in messages:
        content = message.get("content")
        if isinstance(content, str):
            prompt = content
            continue
        if not isinstance(content, Sequence):
            continue
        for part in content:
            if not isinstance(part, Mapping):
                continue
            if part.get("type") == "text":
                prompt = str(part.get("text") or "")
            elif part.get("type") == "image_url":
                image_url = part.get("image_url")
                if isinstance(image_url, Mapping):
                    data_uri = str(image_url.get("url") or "")
                elif isinstance(image_url, str):
                    data_uri = image_url

    if not data_uri.startswith("data:") or ";base64," not in data_uri:
        raise VisualConversionModelError(
            "markitdown-ocr supplied no decodable embedded image"
        )
    header, encoded = data_uri.split(",", 1)
    media_type = header.removeprefix("data:").split(";", 1)[0] or "image/png"
    try:
        image_bytes = base64.b64decode(encoded, validate=True)
    except ValueError as exc:
        raise VisualConversionModelError(
            "markitdown-ocr supplied invalid embedded image data"
        ) from exc
    return prompt, image_bytes, media_type


def _prepare_visual_input(
    image_bytes: bytes,
    media_type: str,
) -> tuple[bytes, str] | None:
    """Bound vision payload size while retaining readable document evidence."""
    if media_type.lower() in LEGACY_VECTOR_MEDIA_TYPES:
        return None
    try:
        with Image.open(BytesIO(image_bytes)) as image:
            image.load()
            if (
                max(image.size) <= VISUAL_MAX_DIMENSION_PIXELS
                and len(image_bytes) <= VISUAL_MAX_PASSTHROUGH_BYTES
                and media_type in {"image/png", "image/jpeg", "image/webp"}
            ):
                return image_bytes, media_type

            image.thumbnail(
                (VISUAL_MAX_DIMENSION_PIXELS, VISUAL_MAX_DIMENSION_PIXELS),
                Image.Resampling.LANCZOS,
            )
            if image.mode in {"RGBA", "LA"} or (
                image.mode == "P" and "transparency" in image.info
            ):
                rgba = image.convert("RGBA")
                flattened = Image.new("RGB", rgba.size, "white")
                flattened.paste(rgba, mask=rgba.getchannel("A"))
                image = flattened
            elif image.mode != "RGB":
                image = image.convert("RGB")

            encoded = BytesIO()
            image.save(
                encoded,
                format="JPEG",
                quality=VISUAL_JPEG_QUALITY,
                optimize=True,
            )
            return encoded.getvalue(), "image/jpeg"
    except Exception as exc:
        raise VisualConversionModelError(
            "Embedded visual content could not be normalized for the vision model"
        ) from exc


def _visual_cache_path(markdown_dir: Path, digest: str) -> Path:
    path = markdown_dir / ".visual-cache" / digest[:2] / f"{digest}.md"
    _assert_generated_path(markdown_dir, path)
    return path


def _read_visual_cache(path: Path) -> str:
    try:
        return _normalize_markdown(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, UnicodeError):
        return ""


def _write_visual_cache(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(
        f".{path.name}.{os.getpid()}.{os.urandom(4).hex()}.tmp"
    )
    try:
        _write_fsynced(temporary, (text + "\n").encode("utf-8"))
        os.replace(temporary, path)
        _fsync_directory(path.parent)
    finally:
        temporary.unlink(missing_ok=True)


def _render_part(
    source: SourceDocument,
    *,
    part_id: str,
    converted_at: datetime,
    anchors: Mapping[str, int | str],
    body: str,
) -> str:
    lines = [
        "---",
        f"document_id: {source.document_id}",
        f"source_path: {_yaml_scalar(source.source_path)}",
        f"source_sha256: {_yaml_scalar(source.source_sha256)}",
        f"part_id: {_yaml_scalar(part_id)}",
        f"converted_at: {_yaml_scalar(converted_at.isoformat())}",
        f"converter_version: {_yaml_scalar(CONVERTER_VERSION)}",
    ]
    for name in ("page", "slide", "sheet", "rows", "section", "segment"):
        if name in anchors:
            lines.append(f"{name}: {_yaml_scalar(anchors[name])}")
    lines.extend(["---", "", _normalize_markdown(body), ""])
    return "\n".join(lines)


def _yaml_scalar(value: object) -> str:
    if isinstance(value, int):
        return str(value)
    return json.dumps(str(value), ensure_ascii=False)


def _render_table_batch(
    batch: Sequence[tuple[int, Sequence[Any]]],
    *,
    sheet_name: str | None,
) -> PartDraft | None:
    nonempty = [
        (row_number, values)
        for row_number, values in batch
        if any(value is not None and value != "" for value in values)
    ]
    if not nonempty:
        return None

    first_row = batch[0][0]
    last_row = batch[-1][0]
    column_count = max(len(values) for _, values in nonempty)
    headings = ["Row", *(get_column_letter(index) for index in range(1, column_count + 1))]
    table_lines = [
        "| " + " | ".join(headings) + " |",
        "| " + " | ".join("---" for _ in headings) + " |",
    ]
    for row_number, values in nonempty:
        padded = [*values, *(None for _ in range(column_count - len(values)))]
        cells = [str(row_number), *(_markdown_cell(value) for value in padded)]
        table_lines.append("| " + " | ".join(cells) + " |")

    anchors: dict[str, int | str] = {"rows": f"{first_row}-{last_row}"}
    heading = ""
    if sheet_name is not None:
        anchors["sheet"] = sheet_name
        heading = f"## Sheet: {sheet_name}\n\n"
    body = heading + f"Rows {first_row}-{last_row}\n\n" + "\n".join(table_lines)
    return PartDraft(body=body, anchors=anchors)


def _markdown_cell(value: Any) -> str:
    if value is None:
        text = ""
    elif isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    elif isinstance(value, datetime):
        text = value.isoformat(sep=" ")
    elif isinstance(value, (date, time)):
        text = value.isoformat()
    elif isinstance(value, float):
        if math.isnan(value):
            text = "NaN"
        elif math.isinf(value):
            text = "Infinity" if value > 0 else "-Infinity"
        else:
            text = str(value)
    elif isinstance(value, Decimal):
        text = format(value, "f")
    else:
        text = str(value)
    return (
        text.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", "<br>")
        .replace("\r", "<br>")
        .replace("\n", "<br>")
    )


def _xls_cell_value(cell: Any, workbook: Any) -> Any:
    import xlrd

    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, workbook.datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return f"#ERROR:{xlrd.error_text_from_code.get(cell.value, cell.value)}"
    return cell.value


def _split_anchored_blocks(
    markdown: str,
    pattern: re.Pattern[str],
    *,
    anchor_name: str,
) -> list[PartDraft]:
    matches = list(pattern.finditer(markdown))
    if not matches:
        return []
    drafts = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(markdown)
        body = markdown[match.start() : end].strip()
        drafts.append(
            PartDraft(
                body=body,
                anchors={anchor_name: int(match.group(1))},
            )
        )
    return drafts


def _has_meaningful_pdf_text(body: str) -> bool:
    without_heading = re.sub(
        r"(?im)^##\s+Page\s+\d+\s*$",
        "",
        body,
    )
    meaningful = re.findall(r"[\w\u3400-\u9fff]", without_heading)
    return len(meaningful) >= PDF_TEXT_MIN_CHARACTERS


def _chunk_text(text: str, limit: int) -> list[str]:
    if limit <= 0:
        raise ValueError("Text chunk size must be positive")
    if len(text) <= limit:
        return [text]

    chunks: list[str] = []
    remaining = text
    while len(remaining) > limit:
        split_at = remaining.rfind("\n\n", 0, limit)
        if split_at < limit // 2:
            split_at = remaining.rfind("\n", 0, limit)
        if split_at < limit // 2:
            split_at = limit
        chunks.append(remaining[:split_at].strip())
        remaining = remaining[split_at:].lstrip()
    if remaining:
        chunks.append(remaining.strip())
    return chunks


def _chunk_markdown(text: str, limit: int) -> list[str]:
    """Chunk text while keeping a whole-document fenced block valid per part."""
    outer_fence = re.fullmatch(r"(```[^\n]*\n)(.*)(\n```)\s*", text, re.DOTALL)
    if outer_fence is None or len(text) <= limit:
        return _chunk_text(text, limit)
    opening, body, closing = outer_fence.groups()
    body_limit = max(1, limit - len(opening) - len(closing))
    return [
        f"{opening}{chunk}{closing}"
        for chunk in _chunk_text(body, body_limit)
    ]


def _read_text(path: Path) -> str:
    encoding = _detect_text_encoding(path)
    try:
        return path.read_text(encoding=encoding)
    except (UnicodeError, OSError) as exc:
        raise DocumentConversionError("Text source could not be decoded") from exc


def _xlsx_contains_embedded_media(path: Path) -> bool:
    try:
        with zipfile.ZipFile(path) as archive:
            return any(name.startswith("xl/media/") for name in archive.namelist())
    except (OSError, zipfile.BadZipFile) as exc:
        raise DocumentConversionError(
            "XLSX package could not be inspected for embedded media"
        ) from exc


def _detect_text_encoding(path: Path) -> str:
    sample_size = 128 * 1024
    try:
        with path.open("rb") as source:
            sample = source.read(sample_size)
    except OSError as exc:
        raise DocumentConversionError("Source file could not be read") from exc
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        sample.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        from charset_normalizer import from_bytes

        best = from_bytes(sample).best()
        if best is None or best.encoding is None:
            raise DocumentConversionError("Text encoding could not be determined")
        return best.encoding


def _normalize_markdown(markdown: str) -> str:
    return markdown.replace("\r\n", "\n").replace("\r", "\n").strip()


def _write_fsynced(path: Path, data: bytes) -> None:
    with path.open("xb") as artifact_file:
        artifact_file.write(data)
        artifact_file.flush()
        os.fsync(artifact_file.fileno())


def _fsync_directory(path: Path) -> None:
    try:
        descriptor = os.open(path, os.O_RDONLY)
    except OSError:
        return
    try:
        os.fsync(descriptor)
    except OSError:
        pass
    finally:
        os.close(descriptor)


def _assert_generated_path(root: Path, candidate: Path) -> None:
    resolved_root = root.resolve(strict=False)
    try:
        candidate.resolve(strict=False).relative_to(resolved_root)
    except ValueError as exc:
        raise ArtifactPublishError("Generated artifact path escapes its root") from exc
